from typing import Optional 
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
import os
import re
import json
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from typing import List

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, ListFlowable, ListItem
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.units import inch


def get_next_filename(base_name="brd", ext="pdf", folder="reports"):
    os.makedirs(folder, exist_ok=True)
    existing_files = os.listdir(folder)
    
    version_pattern = re.compile(rf"{re.escape(base_name)}_v(\d+)\.{ext}")
    versions = [
        int(match.group(1)) 
        for f in existing_files 
        if (match := version_pattern.match(f))
    ]
    
    next_version = max(versions, default=0) + 1
    return os.path.join(folder, f"{base_name}_v{next_version}.{ext}")

def save_logs(pdf_filename, brd_text):
    root_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    folder_name = os.path.join(root_dir, 'logs')
    os.makedirs(folder_name,exist_ok=True)
    log_file = os.path.join(folder_name,"brd_log.json")
    log_entry = {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "pdf": pdf_filename,
        "brd_text": brd_text
    }

    # Load existing logs
    if os.path.exists(log_file):
        with open(log_file, "r", encoding="utf-8") as f:
            logs = json.load(f)
    else:
        logs = []

    logs.append(log_entry)

    # Save updated logs
    with open(log_file, "w", encoding="utf-8") as f:
        json.dump(logs, f, indent=2)
    print("The logs saved sucessfully in :",log_file)


def markdown_to_paragraphs(markdown_text: str, styles) -> list:
    """
    Converts basic markdown text to a list of reportlab flowables (Paragraph, Spacer, List).
    """
    flowables = []
    lines = markdown_text.strip().split('\n')
    bullet_items = []
    numbered_items = []

    for line in lines:
        line = line.strip()

        # Handle headers
        if line.startswith('####'):
            flowables.append(Paragraph(f"<b>{line[4:].strip()}</b>", styles['MyHeading4']))
        elif line.startswith('###'):
            flowables.append(Paragraph(f"<b>{line[3:].strip()}</b>", styles['MyHeading3']))
        elif line.startswith('##'):
            flowables.append(Paragraph(f"<b>{line[2:].strip()}</b>", styles['MyHeading2']))
        elif line.startswith('#'):
            flowables.append(Paragraph(f"<b>{line[1:].strip()}</b>", styles['MyHeading1']))

        # Handle bullet list
        elif line.startswith('- '):
            bullet_items.append(Paragraph(markdown_inline_to_html(line[2:].strip()), styles['Normal']))
        elif re.match(r'^\d+\.\s+', line):
            numbered_items.append(Paragraph(markdown_inline_to_html(re.sub(r'^\d+\.\s+', '', line)), styles['Normal']))

        # Empty line
        elif line == "":
            # Flush bullet list if any
            if bullet_items:
                flowables.append(ListFlowable(
                    [ListItem(b, leftIndent=20) for b in bullet_items],
                    bulletType='bullet', start='-', leftIndent=20))
                bullet_items = []
            if numbered_items:
                flowables.append(ListFlowable(
                    [ListItem(n, leftIndent=20) for n in numbered_items],
                    bulletType='1', leftIndent=20))
                numbered_items = []
            flowables.append(Spacer(1, 12))

        # Normal paragraph
        else:
            flowables.append(Paragraph(markdown_inline_to_html(line), styles['Normal']))

    # Flush remaining list items
    if bullet_items:
        flowables.append(ListFlowable([ListItem(b, leftIndent=20) for b in bullet_items],
                                      bulletType='bullet', start='-', leftIndent=20))
    if numbered_items:
        flowables.append(ListFlowable([ListItem(n, leftIndent=20) for n in numbered_items],
                                      bulletType='1', leftIndent=20))

    return flowables


def markdown_inline_to_html(text: str) -> str:
    """
    Convert inline markdown (**bold**, *italic*, __underline__) to HTML-style for Paragraph.
    """
    text = re.sub(r'\*\*(.*?)\*\*', r'<b>\1</b>', text)
    text = re.sub(r'\*(.*?)\*', r'<i>\1</i>', text)
    text = re.sub(r'__(.*?)__', r'<u>\1</u>', text)
    return text


def save_report(report: str) -> str:
    """
    Converts a BRD (Business Requirement Document) string into a formatted PDF file 
    and saves it with an auto-incremented versioned filename (e.g., brd_v1.pdf, brd_v2.pdf, etc.).

    Args:
        report (str): The full text content of the BRD to be included in the PDF.

    Returns:
        str: The file path of the saved PDF document.
    """
    root_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    text_output_folder = os.path.join(root_dir, 'Text_Output')
    os.makedirs(text_output_folder, exist_ok=True)
    text_filename="brd_doc.txt"
    text_output_path = os.path.join(text_output_folder, text_filename)

    with open(text_output_path,'w') as f:
        f.write(report)

    root_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    output_folder = os.path.join(root_dir, 'Output')
    os.makedirs(output_folder, exist_ok=True)
    filename=get_next_filename(base_name="BRD",ext="pdf",folder=output_folder)
    output_path = os.path.join(output_folder, filename)

    doc = SimpleDocTemplate(output_path, pagesize=A4,
                            rightMargin=50, leftMargin=50,
                            topMargin=50, bottomMargin=50)

    # Styles
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='MyHeading1', parent=styles['Heading1'], fontSize=18, spaceAfter=12))
    styles.add(ParagraphStyle(name='MyHeading2', parent=styles['Heading2'], fontSize=16, spaceAfter=10))
    styles.add(ParagraphStyle(name='MyHeading3', parent=styles['Heading3'], fontSize=14, spaceAfter=8))
    styles.add(ParagraphStyle(name='MyHeading4', parent=styles['Heading4'], fontSize=12, spaceAfter=6))
    styles.add(ParagraphStyle(name='MyNormal',    parent=styles['Normal'],  fontSize=10, spaceAfter=6, alignment=TA_LEFT))

    story = markdown_to_paragraphs(report, styles)
    doc.build(story)

    print("Report saved in path:", output_path)
    save_logs(filename,report)
    return output_path


def save_user_manual(report: str) -> str:
    """
    Converts a user manual report string into a formatted PDF file 
    and saves it with an auto-incremented versioned filename (e.g., user_manual_v1.pdf, user_manual_v2.pdf, etc.).

    Args:
        report (str): The full text content of the user manual to be included in the PDF.

    Returns:
        str: The file path of the saved PDF document.
    """
    # root_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    # text_output_folder = os.path.join(root_dir, 'Text_Output')
    # os.makedirs(text_output_folder, exist_ok=True)
    # text_filename="brd_doc.txt"
    # text_output_path = os.path.join(text_output_folder, text_filename)

    # with open(text_output_path,'w') as f:
    #     f.write(report)

    root_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    output_folder = os.path.join(root_dir, 'User_Manual')
    os.makedirs(output_folder, exist_ok=True)
    filename=get_next_filename(base_name="user_manual",ext="pdf",folder=output_folder)
    output_path = os.path.join(output_folder, filename)

    doc = SimpleDocTemplate(output_path, pagesize=A4,
                            rightMargin=50, leftMargin=50,
                            topMargin=50, bottomMargin=50)

    # Styles
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='MyHeading1', parent=styles['Heading1'], fontSize=18, spaceAfter=12))
    styles.add(ParagraphStyle(name='MyHeading2', parent=styles['Heading2'], fontSize=16, spaceAfter=10))
    styles.add(ParagraphStyle(name='MyHeading3', parent=styles['Heading3'], fontSize=14, spaceAfter=8))
    styles.add(ParagraphStyle(name='MyHeading4', parent=styles['Heading4'], fontSize=12, spaceAfter=6))
    styles.add(ParagraphStyle(name='MyNormal',    parent=styles['Normal'],  fontSize=10, spaceAfter=6, alignment=TA_LEFT))

    story = markdown_to_paragraphs(report, styles)
    doc.build(story)

    print("Report saved in path:", output_path)
    save_logs(filename,report)
    return output_path

def save_usecase_acceptance_criteria(report: str) -> str:
    """
    Converts a Use Case and Acceptance Criteria string into a formatted PDF file 
    and saves it with an auto-incremented versioned filename (e.g., use_case_v1.pdf, use_case_v2.pdf, etc.).

    Args:
        report (str): The full text content of the Use Case and Acceptance Criteria to be included in the PDF.

    Returns:
        str: The file path of the saved PDF document.
    """
    # root_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    # text_output_folder = os.path.join(root_dir, 'Text_Output')
    # os.makedirs(text_output_folder, exist_ok=True)
    # text_filename="brd_doc.txt"
    # text_output_path = os.path.join(text_output_folder, text_filename)

    # with open(text_output_path,'w') as f:
    #     f.write(report)

    root_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    output_folder = os.path.join(root_dir, 'Usecase')
    os.makedirs(output_folder, exist_ok=True)
    filename=get_next_filename(base_name="Usecase",ext="pdf",folder=output_folder)
    output_path = os.path.join(output_folder, filename)

    doc = SimpleDocTemplate(output_path, pagesize=A4,
                            rightMargin=50, leftMargin=50,
                            topMargin=50, bottomMargin=50)

    # Styles
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='MyHeading1', parent=styles['Heading1'], fontSize=18, spaceAfter=12))
    styles.add(ParagraphStyle(name='MyHeading2', parent=styles['Heading2'], fontSize=16, spaceAfter=10))
    styles.add(ParagraphStyle(name='MyHeading3', parent=styles['Heading3'], fontSize=14, spaceAfter=8))
    styles.add(ParagraphStyle(name='MyHeading4', parent=styles['Heading4'], fontSize=12, spaceAfter=6))
    styles.add(ParagraphStyle(name='MyNormal',    parent=styles['Normal'],  fontSize=10, spaceAfter=6, alignment=TA_LEFT))

    story = markdown_to_paragraphs(report, styles)
    doc.build(story)

    print("Report saved in path:", output_path)
    save_logs(filename,report)
    return output_path

# old function without status column

# def save_task_chart(tasks:str):
#     tasks=json.loads(tasks)

#     # 2. Convert dates and get full date range
#     for t in tasks:
#         t["start_date"] = pd.to_datetime(t["start_date"])
#         t["end_date"] = pd.to_datetime(t["end_date"])

#     start_all = min(t["start_date"] for t in tasks)
#     end_all = max(t["end_date"] for t in tasks)
#     all_dates = pd.date_range(start_all, end_all)

#     # 3. Create initial Excel file
#     root_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
#     output_folder = os.path.join(root_dir, 'Task_Chart')
#     os.makedirs(output_folder, exist_ok=True)
#     excel_file = "task_chart.xlsx"
#     excel_file = os.path.join(output_folder,excel_file)
#     columns = ["Task", "Start Date", "End Date"] + [date.strftime("%Y-%m-%d") for date in all_dates]
#     rows = []
#     for t in tasks:
#         row = {
#             "Task": t["task"],
#             "Start Date": t["start_date"].strftime("%Y-%m-%d"),
#             "End Date": t["end_date"].strftime("%Y-%m-%d")
#         }
#         rows.append(row)

#     df = pd.DataFrame(rows, columns=columns)
#     df.to_excel(excel_file, index=False)

#     # 4. Color Gantt bars with different colors & insert task names
#     wb = openpyxl.load_workbook(excel_file)
#     ws = wb.active

#     # Color palette
#     task_colors = [
#         "4F81BD", "C0504D", "9BBB59", "8064A2",
#         "F79646", "2C4D75", "00B0F0", "92D050"
#     ]

#     font = Font(color="FFFFFF", bold=True)
#     align = Alignment(horizontal="center", vertical="center")

#     for row_idx, task, in enumerate(tasks, start=2):
#         start = task["start_date"]
#         end = task["end_date"]
#         duration = (end - start).days + 1

#         start_col_offset = (start - start_all).days + 4  # offset for Task, Start, End
#         end_col_offset = start_col_offset + duration - 1

#         # Create fill for this task
#         color_hex = task_colors[(row_idx - 2) % len(task_colors)]
#         fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

#         # Merge cells for task duration
#         ws.merge_cells(
#             start_row=row_idx,
#             start_column=start_col_offset,
#             end_row=row_idx,
#             end_column=end_col_offset
#         )

#         # Write task name inside the merged cell
#         cell = ws.cell(row=row_idx, column=start_col_offset)
#         cell.value = task["task"]
#         cell.fill = fill
#         cell.font = font
#         cell.alignment = align

#         # Apply fill to all cells in the merged range
#         for col in range(start_col_offset, end_col_offset + 1):
#             ws.cell(row=row_idx, column=col).fill = fill

#     # Auto-size columns
#     for col in ws.columns:
#         max_length = 0
#         column = col[0].column_letter
#         for cell in col:
#             if cell.value:
#                 max_length = max(max_length, len(str(cell.value)))
#         ws.column_dimensions[column].width = max_length + 2

#     wb.save(excel_file)
#     print(f"Gantt chart saved to: {excel_file}")
#     return excel_file


def save_task_chart(tasks: str):
    tasks = json.loads(tasks)

    # Convert dates
    for t in tasks:
        t["start_date"] = pd.to_datetime(t["start_date"])
        t["end_date"] = pd.to_datetime(t["end_date"])

    # Get overall date range
    start_all = min(t["start_date"] for t in tasks)
    end_all = max(t["end_date"] for t in tasks)
    all_dates = pd.date_range(start_all, end_all)

    # # Prepare output folder
    root_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    output_folder = os.path.join(root_dir, 'Task_Chart')
    os.makedirs(output_folder, exist_ok=True)
    excel_file = "task_chart.xlsx"
    excel_file = os.path.join(output_folder,excel_file)

    # Define columns including the new Status column
    columns = ["Task", "Start Date", "End Date", "Status"] + [date.strftime("%Y-%m-%d") for date in all_dates]

    # Create DataFrame rows
    rows = []
    for t in tasks:
        row = {
            "Task": t["task"],
            "Start Date": t["start_date"].strftime("%Y-%m-%d"),
            "End Date": t["end_date"].strftime("%Y-%m-%d"),
            "Status": ""  # Empty to be filled by dropdown
        }
        rows.append(row)

    df = pd.DataFrame(rows, columns=columns)
    df.to_excel(excel_file, index=False)

    # Load workbook to apply formatting
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # Define colors
    task_colors = [
        "4F81BD", "C0504D", "9BBB59", "8064A2",
        "F79646", "2C4D75", "00B0F0", "92D050"
    ]

    font = Font(color="FFFFFF", bold=True)
    align = Alignment(horizontal="center", vertical="center")

    # Apply Gantt colors
    for row_idx, task in enumerate(tasks, start=2):
        start = task["start_date"]
        end = task["end_date"]
        duration = (end - start).days + 1

        # Offset: Task, Start, End, Status = 4 columns
        start_col_offset = (start - start_all).days + 5
        end_col_offset = start_col_offset + duration - 1

        color_hex = task_colors[(row_idx - 2) % len(task_colors)]
        fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

        # Merge cells for the task
        ws.merge_cells(
            start_row=row_idx,
            start_column=start_col_offset,
            end_row=row_idx,
            end_column=end_col_offset
        )

        cell = ws.cell(row=row_idx, column=start_col_offset)
        cell.value = task["task"]
        cell.fill = fill
        cell.font = font
        cell.alignment = align

        # Fill each cell in the range
        for col in range(start_col_offset, end_col_offset + 1):
            ws.cell(row=row_idx, column=col).fill = fill

    # Add dropdown for Status column (D column, i.e., 4th column)
    dv = DataValidation(type="list", formula1='"In Progress,On Hold,Completed"', allow_blank=True)
    status_range = f"D2:D{len(tasks)+1}"
    dv.add(status_range)
    ws.add_data_validation(dv)

    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    dv = DataValidation(type="list", formula1='"In Progress,On Hold,Completed"', allow_blank=True)
    ws.add_data_validation(dv)
    status_col = 4  # Column D
    for row_idx in range(2, len(tasks) + 2):
        cell = ws.cell(row=row_idx, column=status_col)
        dv.add(cell)

    # Conditional formatting rules
    status_column_letter = get_column_letter(status_col)

    # Rule for "In Progress" - Yellow
    ws.conditional_formatting.add(
        f"{status_column_letter}2:{status_column_letter}{len(tasks)+1}",
        FormulaRule(
            formula=[f'${status_column_letter}2="In Progress"'],
            fill=PatternFill(start_color="F8FF00", end_color="F8FF00", fill_type="solid")
        )
    )

    # Rule for "On Hold" - Blue
    ws.conditional_formatting.add(
        f"{status_column_letter}2:{status_column_letter}{len(tasks)+1}",
        FormulaRule(
            formula=[f'${status_column_letter}2="On Hold"'],
            fill=PatternFill(start_color="00F7FF", end_color="00F7FF", fill_type="solid")
        )
    )

    # Rule for "Completed" - Green
    ws.conditional_formatting.add(
        f"{status_column_letter}2:{status_column_letter}{len(tasks)+1}",
        FormulaRule(
            formula=[f'${status_column_letter}2="Completed"'],
            fill=PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        )
    )

    wb.save(excel_file)
    print(f"Gantt chart saved to: {excel_file}")
    return excel_file


if __name__ == "__main__":
#     tasks = [
#     {"task": "Research competitors", "start_date": "2025-08-01", "end_date": "2025-08-03"},
#     {"task": "Create prototype", "start_date": "2025-08-04", "end_date": "2025-08-08"},
#     {"task": "Write documentation", "start_date": "2025-08-02", "end_date": "2025-08-06"},
#     {"task": "Write journal", "start_date": "2025-08-03", "end_date": "2025-08-20"}

# ]
    tasks='[{"task": "Requirement Gathering", "start_date": "2025-08-01", "end_date": "2025-08-03"}, {"task": "Stakeholder Meeting", "start_date": "2025-08-02", "end_date": "2025-08-02"}, {"task": "Feasibility Analysis", "start_date": "2025-08-03", "end_date": "2025-08-05"}, {"task": "Project Planning", "start_date": "2025-08-04", "end_date": "2025-08-06"}, {"task": "UI/UX Wireframe Design", "start_date": "2025-08-05", "end_date": "2025-08-08"}, {"task": "Database Schema Design", "start_date": "2025-08-06", "end_date": "2025-08-09"}, {"task": "Backend API Setup", "start_date": "2025-08-08", "end_date": "2025-08-12"}, {"task": "Frontend Layout Implementation", "start_date": "2025-08-09", "end_date": "2025-08-13"}, {"task": "Authentication Module Integration", "start_date": "2025-08-12", "end_date": "2025-08-14"}, {"task": "Dashboard Development", "start_date": "2025-08-13", "end_date": "2025-08-17"}, {"task": "Notification System Setup", "start_date": "2025-08-14", "end_date": "2025-08-15"}, {"task": "QA Test Case Preparation", "start_date": "2025-08-16", "end_date": "2025-08-17"}, {"task": "Unit Testing", "start_date": "2025-08-17", "end_date": "2025-08-19"}, {"task": "Bug Fixing Phase 1", "start_date": "2025-08-18", "end_date": "2025-08-20"}, {"task": "User Feedback Collection", "start_date": "2025-08-20", "end_date": "2025-08-21"}, {"task": "Performance Optimization", "start_date": "2025-08-21", "end_date": "2025-08-23"}, {"task": "Final QA Review", "start_date": "2025-08-24", "end_date": "2025-08-25"}, {"task": "Client Demo Preparation", "start_date": "2025-08-26", "end_date": "2025-08-27"}, {"task": "Client Presentation", "start_date": "2025-08-28", "end_date": "2025-08-28"}, {"task": "Project Closure & Documentation", "start_date": "2025-08-29", "end_date": "2025-08-31"}]'
    save_task_chart(tasks)